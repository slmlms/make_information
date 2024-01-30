# 导入所需库
import comtypes
import openpyxl
from comtypes.gen import Excel
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 定义函数add_row，用于向指定Excel文件的每个工作表添加一行数据，并设置相关格式
def add_row(workbook_path):
    # 定义要写入的数据字符串（包含四个单位名称）
    data_blank = '          '
    data = f'分包单位：{data_blank}总包单位：{data_blank}监理单位：{data_blank}建设单位：{data_blank}'

    try:
        # 加载指定路径的Excel工作簿，仅读取纯数据
        wb = openpyxl.load_workbook(workbook_path, data_only=True)

        # 遍历工作簿中的所有工作表
        for sheet_name in wb.sheetnames:
            # 获取当前工作表对象
            ws = wb[sheet_name]

            # 计算新的行索引（即最大行数加1）
            rows = ws.max_row + 1

            # 在A列的新行中写入数据
            ws.cell(row=rows, column=1).value = data

            # 合并新行的A到D列单元格
            ws.merge_cells(start_row=rows, end_row=rows, start_column=1, end_column=6)

            # 获取合并后的单元格，并设置其对齐方式为水平居中、垂直居中
            merged_cell = ws.cell(row=rows, column=1)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置新增行的高度为25
            ws.row_dimensions[rows].height = 25

            # 计算当前工作表的最大列并设置打印区域（从A1到最大列的最大行）
            max_column = ws.max_column
            print_area = f"A1:{get_column_letter(max_column)}{rows}"
            ws.print_area = print_area

        # 使用with语句确保工作簿在操作完成后被正确关闭并保存
        with wb:
            wb.save(workbook_path)

    except Exception as e:
        print(f"An error occurred: {e}")

# 调用add_row函数，传入待处理的Excel文件路径
workbook_path = "D:\\Jobs\\卡莫亚\\签证联系单\\联系单\\电气工程量\\选矿电气工程量1.xlsx"
add_row(workbook_path)


# 定义函数excel_sheets_to_pdf，将Excel文件的所有工作表导出为PDF格式
def excel_sheets_to_pdf(excel_path, output_base_name):
    # 创建Excel应用对象
    excel_app = comtypes.client.CreateObject('Excel.Application')

    try:
        # 打开指定路径的Excel文件
        wb = excel_app.Workbooks.Open(Filename=excel_path)

        # 遍历工作簿中的所有工作表及其索引
        for sheet_index, sheet in enumerate(wb.Sheets):
            # 根据索引和工作表名称设置PDF输出路径
            pdf_output = f"{output_base_name}_{sheet.Name}.pdf"

            # 使用Excel的ExportAsFixedFormat方法将当前工作表导出为PDF
            sheet.ExportAsFixedFormat(Type=Excel.xlTypePDF,
                                      Filename=pdf_output,
                                      Quality=Excel.xlQualityStandard,
                                      IncludeDocProperties=True,
                                      IgnorePrintAreas=False,
                                      OpenAfterPublish=False)

        # 保存并关闭工作簿（不保存更改，因为我们在导出为PDF时已经包含了需要的信息）
        wb.Close(SaveChanges=False)

    finally:
        # 确保退出Excel进程
        excel_app.Quit()

# 调用excel_sheets_to_pdf函数，传入待处理的Excel文件路径及PDF输出的基础名称
excel_sheets_to_pdf(r"D:\Jobs\卡莫亚\签证联系单\联系单\电气工程量\选矿电气工程量1.xlsx", "D:\Jobs\卡莫亚\签证联系单\联系单\电气工程量\电气工程量\\")
