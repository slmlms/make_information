# 导入所需模块
import logging
import os
import random

import openpyxl
import pandas as pd
from docxtpl import DocxTemplate
from openpyxl.reader.excel import load_workbook
from tqdm import tqdm

logging.basicConfig(level=logging.INFO)  # 设置日志级别为INFO
# 文件路径
Jianyanpi_data_path = "D:\Jobs\卡莫亚\检验批及分项\钢结构检验批.xlsx"
Jianyanpi_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\钢结构安装\检验批\\"
Jianyanpi_save_path = "D:\Jobs\卡莫亚\检验批及分项\钢结构检验批生成\\"
Fenxiang_tamepate_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\钢结构安装\分项质量检查验收记录.docx"
Fenxiang_Baoyan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\钢结构安装\分项报验申请表.docx"
Fenbu_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\钢结构安装\分部工程质量检验评定记录.docx"
Fenbu_Baoyan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\钢结构安装\分部报验申请表.docx"


# 从Excel文件中读取指定工作表到DataFrame中
def read_excel_to_dataframe(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df


# 获取模板文件路径
def get_template_file_path(sheet, row):
    """
    根据给定的工作表和行索引，获取模板文件路径。

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): 工作表对象。
        row (int): 行索引。

    Returns:
        str: 模板文件的完整路径。
    """
    cell_value = sheet.cell(row=row, column=4).value
    if "屋面瓦" in cell_value or "墙面瓦" in cell_value:
        template_name = sheet.cell(row=row, column=3).value + "检验批质量检验评定记录-" + cell_value + ".docx"
        template_file_path = Jianyanpi_template_path + template_name
    else:
        template_file_path = Jianyanpi_template_path + sheet.cell(row=row, column=3).value + "检验批质量检验评定记录.docx"
    return template_file_path


def extract_filename_from_path(file_path):
    """
    从文件路径中提取文件名（不包含扩展名）。

    Args:
        file_path (str): 文件路径。

    Returns:
        str: 文件名（不包含扩展名）。
    """
    base_name = os.path.basename(file_path)
    filename_without_extension, extension = os.path.splitext(base_name)
    return filename_without_extension


def jyp_wumianwa(data):
    """
    生成屋面挖检验批数据，包括檐口与屋脊的平行度、压型金属板波纹对屋脊的垂直度、
    檐口相邻两块压型金属板端部错位、压型金属板卷边板件最大波浪高。

    Args:
        data (dict): 包含检验批数据的字典。

    Returns:
        dict: 更新后的包含检验批数据的字典。
    """
    for i in range(1, 11):
        data["pingxingdu" + str(i)] = random.randint(1, 3)  # 檐口与屋脊的平行度
        data["chuizhidu" + str(i)] = random.randint(5, 9)  # 压型金属板波纹对屋脊的垂直度
        data["cuowei" + str(i)] = random.randint(1, 4)  # 檐口相邻两块压型金属板端部错位
        data["bolanggao" + str(i)] = random.randint(1, 3)  # 压型金属板卷边板件最大波浪高
    return data


def jyp_qiangmianwa(data):
    """
    生成墙面挖检验批数据，包括墙板波纹线垂直度、墙板包角板的垂直度、
    相邻两块压型金属板的下端错位。

    Args:
        data (dict): 包含检验批数据的字典。

    Returns:
        dict: 更新后的包含检验批数据的字典。
    """
    for i in range(1, 11):
        data["bowenxian" + str(i)] = random.randint(5, 9)  # 墙板波纹线垂直度
        data["chuizhidu" + str(i)] = random.randint(5, 9)  # 墙板包角板的垂直度
        data["cuowei" + str(i)] = random.randint(1, 4)  # 相邻两块压型金属板的下端错位
    return data


def jyp_gangjiegoufangfu(data):
    """
    生成钢结构防腐检验批数据，包括室外油漆、室内油漆。

    Args:
        data (dict): 包含检验批数据的字典。

    Returns:
        dict: 更新后的包含检验批数据的字典。
    """
    for i in range(1, 11):
        data["shiwai" + str(i)] = random.randint(2, 20)  # 室外油漆
        data["shinei" + str(i)] = random.randint(2, 20)  # 室内油漆
    return data


def jyp_gangjiegouhanjie(data):
    """
    生成钢结构焊接检验批数据，包括对接等要求熔透的焊缝、吊车梁或类似构件的腹板与上翼缘连接处、
    对接焊缝余高、对接焊缝错边、焊脚尺寸、角焊缝余高。

    Args:
        data (dict): 包含检验批数据的字典。

    Returns:
        dict: 更新后的包含检验批数据的字典。
    """
    for i in range(1, 11):
        data["rongtouhanfeng" + str(i)] = random.randint(0, 3)  # 对接等要求熔透的焊缝
        data["diaocheliang" + str(i)] = random.randint(0, 3)  # 吊车梁或类似构件的腹板与上翼缘连接处
        data["hanfengyugao" + str(i)] = random.randint(0, 3)  # 对接焊缝余高
        data["hanfengcuobian" + str(i)] = random.randint(10, 50) / 100  # 对接焊缝错边
        data["hanjiaochicun" + str(i)] = random.randint(0, 2)  # 焊脚尺寸
        data["jiaohanfengyugao" + str(i)] = random.randint(0, 2)  # 角焊缝余高
    return data


def jyp_dancenggangjiegou(data):
    """
    生成钢结构焊接检验批数据，包括对接等要求熔透的焊缝、吊车梁或类似构件的腹板与上翼缘连接处、
    对接焊缝余高、对接焊缝错边、焊脚尺寸、角焊缝余高。

    Args:
        data (dict): 包含检验批数据的字典。

    Returns:
        dict: 更新后的包含检验批数据的字典。
    """
    for i in range(1, 11):
        data["kuazhongchuizhidu" + str(i)] = random.randint(2, 8)  # 钢屋架、桁架、梁及受压杆跨中垂直度
        data["cexiangwanqu" + str(i)] = random.randint(2, 8)  # 钢屋架、桁架、梁及受压杆侧向弯曲
        data["zhengtichuizhidu" + str(i)] = random.randint(2, 8)  # 主体结构整体垂直度
        data["pingmianwanqudu" + str(i)] = random.randint(2, 8)  # 主体结构整体平面弯曲度
        data["zhouxianpianyi" + str(i)] = random.randint(1, 3)  # 柱脚底座中心线对定位轴线偏移
        data["jizhundianbiaogao" + str(i)] = random.randint(1, 3)  # 柱基准点标高
        data["wanqushigao" + str(i)] = random.randint(1, 8)  # 弯曲矢高
        data["zhouxianchuizhidu" + str(i)] = random.randint(1, 5)  # 柱轴线垂直度
        data["qiangjiazhouxianpianyi" + str(i)] = random.randint(1, 3)  # 墙架立柱中心线对定位轴线的偏移
        data["hengjiachuizhidu" + str(i)] = random.randint(1, 6)  # 抗风桁架的垂直度
        data["lintiaojianju" + str(i)] = random.randint(1, 3)  # 檩条、墙梁的间距
        data["pingtaigaodu" + str(i)] = random.randint(1, 9)  # 平台高度
        data["pingtailiangshuipingdu" + str(i)] = random.randint(5, 9)  # 平台梁水平度
        data["cexiangwanqu" + str(i)] = random.randint(1, 6)  # 承重平台梁侧向弯曲
        data["pingtailiangchuizhidu" + str(i)] = random.randint(1, 4)  # 承重平台梁垂直度
        data["langangaodu" + str(i)] = random.randint(1, 8)  # 栏杆高度
        data["liganjianju" + str(i)] = random.randint(1, 8)  # 栏杆立柱间距
        data["hanfengzudui" + str(i)] = random.randint(1, 2)  # 现场焊缝组对间隙

    return data


# 生成检查单函数
def generate_inspection_batch(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        logging.info("检验批")
        wb = load_workbook(Jianyanpi_data_path)
        logging.info("成功加载Excel文件: %s", Jianyanpi_data_path)  # 添加日志：成功加载Excel文件
        for sheet in wb:
            # if sheet.title != "尾矿浓密及输送": continue
            for row in tqdm(range(2, sheet.max_row + 1), desc=f'Processing Sheet: {sheet.title}'):
                try:
                    template_file_path = get_template_file_path(sheet, row)
                    template_file = DocxTemplate(template_file_path)
                    data = {
                        "Danweigongchengmingcheng": sheet.cell(row=row, column=1).value,
                        "Fenbugongchengmingcheng": sheet.cell(row=row, column=2).value,
                        "Fenxianggongchengmingcheng": sheet.cell(row=row, column=3).value,
                        "Jianyanchibuwei": sheet.cell(row=row, column=4).value,
                        "Jianyanchirongliang": sheet.cell(row=row, column=5).value
                    }
                    # 生成数据
                    if data["Jianyanchibuwei"] == "屋面瓦":
                        data = jyp_wumianwa(data)
                    elif data["Jianyanchibuwei"] == "墙面瓦":
                        data = jyp_qiangmianwa(data)
                    elif data["Fenxianggongchengmingcheng"] == "钢结构焊接":
                        data = jyp_gangjiegouhanjie(data)
                    elif data["Fenxianggongchengmingcheng"] == "钢结构预拼装":
                        pass
                    elif data["Fenxianggongchengmingcheng"] == "钢结构防腐防锈涂料涂装":
                        data = jyp_gangjiegoufangfu(data)
                    elif data["Fenxianggongchengmingcheng"] == "钢结构单层结构安装":
                        data = jyp_dancenggangjiegou(data)
                    for key in data:
                        if data[key] is None:
                            data[key] = ''
                    save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/" + data[
                        'Fenxianggongchengmingcheng'] + '/'
                    if not os.path.exists(os.path.dirname(save_path)):
                        os.makedirs(os.path.dirname(save_path))
                    template_file.render(data)
                    template_file.save(
                        save_path + extract_filename_from_path(template_file_path) + "-验收部位-" + data[
                            "Jianyanchirongliang"] + data[
                            "Jianyanchibuwei"] + '.docx')
                except FileNotFoundError:
                    logging.error("模板文件 %s 不存在", template_file_path)  # 添加日志：模板文件不存在
                except IndexError:
                    logging.error("在行 %s 的列中找不到值", row)  # 添加日志：列值不存在
                except KeyError:
                    logging.error("数据字典中缺少键（单元格值可能为空）")  # 添加日志：缺少键
                except Exception as e:
                    logging.error("处理Sheet '%s' 行 %s 时发生错误: %s", sheet.title, row, str(e))  # 添加日志：处理行时发生错误
    except FileNotFoundError:
        logging.error("找不到文件: %s", Jianyanpi_data_path)  # 添加日志：文件不存在
    except Exception as e:
        logging.error("加载Excel文件时发生错误: %s", str(e))  # 添加日志：加载文件时发生错误
    finally:
        try:
            wb.close()
        except Exception as e:
            logging.error("关闭工作簿时发生错误: %s", str(e))  # 添加日志：关闭工作簿时发生错误


# 生成分项及报验表
def generate_itemised_project(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        logging.info("分项")
        wb = load_workbook(Jianyanpi_data_path)
        for sheet in wb:
            print(sheet.title)
            try:
                df = read_excel_to_dataframe(Jianyanpi_data_path, sheet.title)
                unique_project_names = df['分项工程名称'].unique()
                for unique_project_name in tqdm(unique_project_names,
                                                desc=f'Processing Projects in Sheet: {sheet.title}'):
                    try:
                        tpl = DocxTemplate(Fenxiang_tamepate_path)
                        tpl_baoyan = DocxTemplate(Fenxiang_Baoyan_template_path)
                        lists = []
                        data = {
                            "Danweigongchengmingcheng": sheet.title,
                            "Fenbugongchengmingcheng": "",
                            "Fenxianggongchengmingcheng": unique_project_name,
                            "size": 0,
                            "list": []
                        }
                        project_name = df[df['分项工程名称'] == unique_project_name]['分部工程名称'].unique()
                        if len(project_name) > 0:
                            data["Fenbugongchengmingcheng"] = project_name[0]
                        part_names = df[df['分项工程名称'] == unique_project_name]['检验批部位'].unique()
                        data["size"] = len(part_names)
                        for i in range(len(part_names)):
                            list = {
                                "Xuhao": str(i + 1).zfill(2),
                                "Jianyanchibuwei": part_names[i],
                                "Shigongdanweipingding": "□优良    □合格",
                                "Jianlidanweipingding": "□优良    □合格"
                            }
                            lists.append(list)
                        data["list"] = lists
                        save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/" + \
                                    data['Fenxianggongchengmingcheng'] + '/'
                        if not os.path.exists(os.path.dirname(save_path)):
                            os.makedirs(os.path.dirname(save_path))
                        tpl.render(data)
                        tpl.save(save_path + "02" + data["Fenxianggongchengmingcheng"] + '分项质量检查验收记录.docx')
                        tpl_baoyan.render(data)
                        tpl_baoyan.save(
                            save_path + "01" + data["Fenxianggongchengmingcheng"] + '分项质量检查验收记录-报验表.docx')
                    except FileNotFoundError as fnfe:
                        print(f"模板文件 {Fenxiang_tamepate_path} 不存在. 错误信息: {fnfe}")
                    except KeyError as ke:
                        print(f"在DataFrame中找不到相应的列名. 错误信息: {ke}")
                    except IndexError as ie:
                        print(f"项目名称或部位名称的索引超出范围. 错误信息: {ie}")
                    except Exception as e:
                        print(f"处理Sheet '{sheet.title}' 时发生错误: {str(e)}")
            except ImportError as ime:
                print(f"未能正确导入或使用read_excel_to_dataframe函数. 错误信息: {ime}")
            except ValueError as ve:
                print(f"无法从Excel文件创建DataFrame. 错误信息: {ve}")
            except Exception as e:
                print(f"处理Sheet '{sheet.title}' 数据时发生错误: {str(e)}")
    except FileNotFoundError as fnfe:
        print(f"找不到Excel文件: {Jianyanpi_data_path}. 错误信息: {fnfe}")
    except openpyxl.exceptions.InvalidFileException as ife:
        print(f"Excel文件 {Jianyanpi_data_path} 格式无效. 错误信息: {ife}")
    except Exception as e:
        print(f"加载或处理Excel文件时发生错误: {str(e)}")


# 生成分部及报验表
def generate_fenbu_project(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        logging.info("分部")
        wb = load_workbook(Jianyanpi_data_path)
        for sheet in wb:
            print(sheet.title)
            df = read_excel_to_dataframe(Jianyanpi_data_path, sheet.title)
            unique_project_names = df['分部工程名称'].unique()
            for unique_project_name in tqdm(unique_project_names, desc=f'Processing Projects in Sheet: {sheet.title}'):
                try:
                    data = {
                        "Danweigongchengmingcheng": sheet.title,
                        "Fenbugongchengmingcheng": unique_project_name,
                        "Fenxiang_size": 0,
                        "list": []
                    }
                    tpl = DocxTemplate(Fenbu_template_path)
                    tpl_Baoyan = DocxTemplate(Fenbu_Baoyan_template_path)
                    part_names = df[df['分部工程名称'] == unique_project_name]['分项工程名称'].unique()
                    data['Fenxiang_size'] = len(part_names)
                    lists = []
                    for i in range(len(part_names)):
                        list = {
                            "Xuhao": str(i + 1),
                            "Fenxianggongchengmingcheng": part_names[i],
                            "count_of_jianyanpi": str(
                                len(df[df['分项工程名称'] == part_names[i]]['检验批部位'].unique())),
                            "Shigongdanweipingding": "□优良    □合格",
                            "Jianlidanweipingding": "□优良    □合格"
                        }
                        lists.append(list)
                    data['list'] = lists
                    save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/"
                    if not os.path.exists(os.path.dirname(save_path)):
                        os.makedirs(os.path.dirname(save_path))
                    tpl.render(data)
                    tpl_Baoyan.render(data)
                    try:
                        tpl.save(save_path + "02" + data['Fenbugongchengmingcheng'] + "分部工程质量检验评定记录.docx")
                        tpl_Baoyan.save(
                            save_path + "01" + data['Fenbugongchengmingcheng'] + "分部工程质量检验评定记录-报验表.docx")
                    except Exception as e:
                        print(f"保存Word文档时发生错误: {e}")
                except Exception as e:
                    print(f"处理{unique_project_name}时发生错误: {e}")
    except FileNotFoundError:
        print(f"无法找到Excel文件: {Jianyanpi_data_path}")
    except Exception as e:
        print(f"处理Excel或生成报告时发生未知错误: {e}")


# 调用方法生成检验批
generate_inspection_batch(Jianyanpi_data_path, Jianyanpi_save_path)
# 生成分项及报验表
# generate_itemised_project(Jianyanpi_data_path, Jianyanpi_save_path)
# # # 生成分部及报验表
# generate_fenbu_project(Jianyanpi_data_path, Jianyanpi_save_path)
