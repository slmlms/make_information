# 导入所需模块
import logging
import os
import random

import openpyxl
import pandas as pd
from docxtpl import DocxTemplate
from openpyxl.reader.excel import load_workbook
from tqdm import tqdm

logging.basicConfig(level=logging.INFO)  # 设置日志级别为INFO
# 文件路径
Jianyanpi_data_path = "D:\Jobs\卡莫亚\检验批及分项\非标设备检验批.xlsx"
Jianyanpi_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\检验批\\"
Jianyanpi_save_path = "D:\Jobs\卡莫亚\检验批及分项\非标设备检验批生成\\"
Fenxiang_tamepate_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\分项质量检查验收记录.docx"
Fenxiang_Baoyan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\分项报验申请表.docx"
Fenbu_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\分部工程质量检验评定记录.docx"
Fenbu_Baoyan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\分部报验申请表.docx"
Manshui_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\检验批\满水试验记录.docx"
Fengguan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\检验批\封罐记录.docx"
Zhenkong_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\检验批\抽真空记录.docx"
Meiyou_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\非标设备检验批模板\检验批\煤油渗透记录.docx"


# 从Excel文件中读取指定工作表到DataFrame中
def read_excel_to_dataframe(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df


# 获取模板文件路径
def get_template_file_path(sheet, row):
    """
    根据给定的工作表和行索引，获取模板文件路径。

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): 工作表对象。
        row (int): 行索引。

    Returns:
        str: 模板文件的完整路径。
    """
    cell_value = sheet.cell(row=row, column=4).value
    if "屋面瓦" in cell_value or "墙面瓦" in cell_value:
        template_name = sheet.cell(row=row, column=3).value + "检验批质量检验评定记录-" + cell_value + ".docx"
        template_file_path = Jianyanpi_template_path + template_name
    else:
        template_file_path = Jianyanpi_template_path + sheet.cell(row=row, column=3).value + "检验批质量检验评定记录.docx"
    return template_file_path


def extract_filename_from_path(file_path):
    """
    从文件路径中提取文件名（不包含扩展名）。

    Args:
        file_path (str): 文件路径。

    Returns:
        str: 文件名（不包含扩展名）。
    """
    base_name = os.path.basename(file_path)
    filename_without_extension, extension = os.path.splitext(base_name)
    return filename_without_extension


def jyp_caoguanguanti(data):
    for i in range(1, 11):
        data["biaogaoA" + str(i)] = random.randint(-5, 10)
        data["biaogaoB" + str(i)] = random.randint(3, 8)
        data["gaocha" + str(i)] = random.randint(1, 5)
        data["diban" + str(i)] = random.randint(1, 8)
        data["bibanshangkou" + str(i)] = random.randint(1, 2)
        data["yuanzhou" + str(i)] = random.randint(1, 2)
        data["banjing" + str(i)] = random.randint(-5, 8)
        data["bibangaodu" + str(i)] = random.randint(3, 8)
        data["bibanchuizhidu" + str(i)] = random.randint(3, 9)
        data["jububianxing" + str(i)] = random.randint(1, 7)
        data["dingbubianxing" + str(i)] = random.randint(1, 5)
        data["lishichuizhidu" + str(i)] = random.randint(1, 5)
        data["zhongxinpiancha" + str(i)] = random.randint(1, 6)
        data["bibanjuli" + str(i)] = random.randint(-2, 4)
        data["falanchuizhidu" + str(i)] = random.randint(1, 3)
    return data


def jyp_rongqihanjie(data):
    for i in range(1, 11):
        data["A" + str(i)] = random.randint(1, 2)
        data["B" + str(i)] = random.randint(1, 2)
        data["C" + str(i)] = random.randint(1, 2)
        data["D" + str(i)] = random.randint(0, 1)
        data["E" + str(i)] = random.randint(1, 2)
    return data


# 生成检查单函数
def generate_inspection_batch(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        logging.info("检验批")
        wb = load_workbook(Jianyanpi_data_path)
        logging.info("成功加载Excel文件: %s", Jianyanpi_data_path)  # 添加日志：成功加载Excel文件
        for sheet in wb:
            # if sheet.title != "尾矿浓密及输送": continue
            i = 3
            for row in tqdm(range(2, sheet.max_row + 1), desc=f'Processing Sheet: {sheet.title}'):
                try:
                    template_file_path = get_template_file_path(sheet, row)
                    template_file = DocxTemplate(template_file_path)

                    data = {
                        "Danweigongchengmingcheng": sheet.cell(row=row, column=1).value,
                        "Fenbugongchengmingcheng": sheet.cell(row=row, column=2).value,
                        "Fenxianggongchengmingcheng": sheet.cell(row=row, column=3).value,
                        "Jianyanchibuwei": sheet.cell(row=row, column=4).value,
                        "Jianyanchirongliang": sheet.cell(row=row, column=5).value,
                        "Riqi": sheet.cell(row=row, column=6).value,
                        "Caoguanguige": sheet.cell(row=row, column=7).value,
                        "Caizhi": sheet.cell(row=row, column=8).value
                    }
                    # 生成数据
                    if data["Fenxianggongchengmingcheng"] == "槽体焊接":
                        data = jyp_rongqihanjie(data)
                    elif data["Fenxianggongchengmingcheng"] == "槽体安装":
                        data = jyp_caoguanguanti(data)
                    for key in data:
                        if data[key] is None:
                            data[key] = ''
                    save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/" + data[
                        'Fenxianggongchengmingcheng'] + '/'
                    if not os.path.exists(os.path.dirname(save_path)):
                        os.makedirs(os.path.dirname(save_path))
                    template_file.render(data)
                    template_file.save(
                        save_path + "{:02}".format(i)+"A" + extract_filename_from_path(template_file_path) + "-验收部位-" +
                        data[
                            "Jianyanchirongliang"] + data[
                            "Jianyanchibuwei"] + '.docx')
                    if "槽" in data["Jianyanchibuwei"] and "槽体焊接" == data["Fenxianggongchengmingcheng"]:
                        # 满水试验记录
                        template_manshui = DocxTemplate(Manshui_template_path)
                        template_manshui.render(data)
                        template_manshui.save(save_path + "{:02}".format(i) + "D" + "满水试验记录-验收部位-" + data[
                            "Jianyanchirongliang"] + data[
                                                  "Jianyanchibuwei"] + '.docx')
                        # 封罐记录
                        template_fengguan = DocxTemplate(Fengguan_template_path)
                        template_fengguan.render(data)
                        template_fengguan.save(save_path + "{:02}".format(i) + "E" + "封罐记录-验收部位-" + data[
                            "Jianyanchirongliang"] + data[
                                                   "Jianyanchibuwei"] + '.docx')
                        # 真空记录
                        template_zhankong = DocxTemplate(Zhenkong_template_path)
                        template_zhankong.render(data)
                        template_zhankong.save(save_path + "{:02}".format(i) + "C" + "抽真空记录-验收部位-" + data[
                            "Jianyanchirongliang"] + data[
                                                   "Jianyanchibuwei"] + '.docx')
                        # 煤油记录
                        template_meiyou = DocxTemplate(Meiyou_template_path)
                        template_meiyou.render(data)
                        template_meiyou.save(save_path + "{:02}".format(i) + "B" + "煤油渗透记录-验收部位-" + data[
                            "Jianyanchirongliang"] + data[
                                                 "Jianyanchibuwei"] + '.docx')
                        i = i + 1
                except FileNotFoundError:
                    logging.error("模板文件 %s 不存在", template_file_path)  # 添加日志：模板文件不存在
                except IndexError:
                    logging.error("在行 %s 的列中找不到值", row)  # 添加日志：列值不存在
                except KeyError:
                    logging.error("数据字典中缺少键（单元格值可能为空）")  # 添加日志：缺少键
                except Exception as e:
                    logging.error("处理Sheet '%s' 行 %s 时发生错误: %s", sheet.title, row, str(e))  # 添加日志：处理行时发生错误
    except FileNotFoundError:
        logging.error("找不到文件: %s", Jianyanpi_data_path)  # 添加日志：文件不存在
    except Exception as e:
        logging.error("加载Excel文件时发生错误: %s", str(e))  # 添加日志：加载文件时发生错误
    finally:
        try:
            wb.close()
        except Exception as e:
            logging.error("关闭工作簿时发生错误: %s", str(e))  # 添加日志：关闭工作簿时发生错误


# 生成分项及报验表
def generate_itemised_project(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        logging.info("分项")
        wb = load_workbook(Jianyanpi_data_path)
        for sheet in wb:
            print(sheet.title)
            try:
                df = read_excel_to_dataframe(Jianyanpi_data_path, sheet.title)
                unique_project_names = df['分项工程名称'].unique()
                for unique_project_name in tqdm(unique_project_names,
                                                desc=f'Processing Projects in Sheet: {sheet.title}'):
                    try:
                        tpl = DocxTemplate(Fenxiang_tamepate_path)
                        tpl_baoyan = DocxTemplate(Fenxiang_Baoyan_template_path)
                        lists = []
                        data = {
                            "Danweigongchengmingcheng": sheet.title,
                            "Fenbugongchengmingcheng": "",
                            "Fenxianggongchengmingcheng": unique_project_name,
                            "size": 0,
                            "list": []
                        }
                        project_name = df[df['分项工程名称'] == unique_project_name]['分部工程名称'].unique()
                        if len(project_name) > 0:
                            data["Fenbugongchengmingcheng"] = project_name[0]
                        part_names = df[df['分项工程名称'] == unique_project_name]['检验批部位'].unique()
                        data["size"] = len(part_names)
                        for i in range(len(part_names)):
                            list = {
                                "Xuhao": str(i + 1).zfill(2),
                                "Jianyanchibuwei": part_names[i],
                                "Shigongdanweipingding": "□优良    □合格",
                                "Jianlidanweipingding": "□优良    □合格"
                            }
                            lists.append(list)
                        data["list"] = lists
                        while len(lists) < 20:
                            lists.append({})
                        save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/" + \
                                    data['Fenxianggongchengmingcheng'] + '/'
                        if not os.path.exists(os.path.dirname(save_path)):
                            os.makedirs(os.path.dirname(save_path))
                        tpl.render(data)
                        tpl.save(save_path + "02" + data["Fenxianggongchengmingcheng"] + '分项质量检查验收记录.docx')
                        tpl_baoyan.render(data)
                        tpl_baoyan.save(
                            save_path + "01" + data["Fenxianggongchengmingcheng"] + '分项质量检查验收记录-报验表.docx')
                    except FileNotFoundError as fnfe:
                        print(f"模板文件 {Fenxiang_tamepate_path} 不存在. 错误信息: {fnfe}")
                    except KeyError as ke:
                        print(f"在DataFrame中找不到相应的列名. 错误信息: {ke}")
                    except IndexError as ie:
                        print(f"项目名称或部位名称的索引超出范围. 错误信息: {ie}")
                    except Exception as e:
                        print(f"处理Sheet '{sheet.title}' 时发生错误: {str(e)}")
            except ImportError as ime:
                print(f"未能正确导入或使用read_excel_to_dataframe函数. 错误信息: {ime}")
            except ValueError as ve:
                print(f"无法从Excel文件创建DataFrame. 错误信息: {ve}")
            except Exception as e:
                print(f"处理Sheet '{sheet.title}' 数据时发生错误: {str(e)}")
    except FileNotFoundError as fnfe:
        print(f"找不到Excel文件: {Jianyanpi_data_path}. 错误信息: {fnfe}")
    except openpyxl.exceptions.InvalidFileException as ife:
        print(f"Excel文件 {Jianyanpi_data_path} 格式无效. 错误信息: {ife}")
    except Exception as e:
        print(f"加载或处理Excel文件时发生错误: {str(e)}")


# 生成分部及报验表
def generate_fenbu_project(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        logging.info("分部")
        wb = load_workbook(Jianyanpi_data_path)
        for sheet in wb:
            print(sheet.title)
            df = read_excel_to_dataframe(Jianyanpi_data_path, sheet.title)
            unique_project_names = df['分部工程名称'].unique()
            for unique_project_name in tqdm(unique_project_names, desc=f'Processing Projects in Sheet: {sheet.title}'):
                try:
                    data = {
                        "Danweigongchengmingcheng": sheet.title,
                        "Fenbugongchengmingcheng": unique_project_name,
                        "Fenxiang_size": 0,
                        "list": []
                    }
                    tpl = DocxTemplate(Fenbu_template_path)
                    tpl_Baoyan = DocxTemplate(Fenbu_Baoyan_template_path)
                    part_names = df[df['分部工程名称'] == unique_project_name]['分项工程名称'].unique()
                    data['Fenxiang_size'] = len(part_names)
                    lists = []
                    for i in range(len(part_names)):
                        list = {
                            "Xuhao": str(i + 1),
                            "Fenxianggongchengmingcheng": part_names[i],
                            "count_of_jianyanpi": str(
                                len(df[df['分项工程名称'] == part_names[i]]['检验批部位'].unique())),
                            "Shigongdanweipingding": "□优良    □合格",
                            "Jianlidanweipingding": "□优良    □合格"
                        }
                        lists.append(list)
                    while len(lists) < 9: lists.append({})
                    data['list'] = lists
                    save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/"
                    if not os.path.exists(os.path.dirname(save_path)):
                        os.makedirs(os.path.dirname(save_path))
                    tpl.render(data)
                    tpl_Baoyan.render(data)
                    try:
                        tpl.save(save_path + "02" + data['Fenbugongchengmingcheng'] + "分部工程质量检验评定记录.docx")
                        tpl_Baoyan.save(
                            save_path + "01" + data['Fenbugongchengmingcheng'] + "分部工程质量检验评定记录-报验表.docx")
                    except Exception as e:
                        print(f"保存Word文档时发生错误: {e}")
                except Exception as e:
                    print(f"处理{unique_project_name}时发生错误: {e}")
    except FileNotFoundError:
        print(f"无法找到Excel文件: {Jianyanpi_data_path}")
    except Exception as e:
        print(f"处理Excel或生成报告时发生未知错误: {e}")


# 调用方法生成检验批
generate_inspection_batch(Jianyanpi_data_path, Jianyanpi_save_path)
# 生成分项及报验表
generate_itemised_project(Jianyanpi_data_path, Jianyanpi_save_path)
# # 生成分部及报验表
generate_fenbu_project(Jianyanpi_data_path, Jianyanpi_save_path)
