import os
import pathlib
import subprocess

import pandas as pd
from PyPDF2 import PdfFileMerger
from openpyxl import load_workbook


# 读取指定Excel的指定Sheet并转换为DataFrame
def read_excel_to_dataframe(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df


# 筛选出“额定功率”字段小于等于7.5的内容，然后根据“子项名称”字段的内容进行分组，返回一个新的DataFrame
def filter_and_group(df):
    filtered_df = df[df['额定功率'] <= 30]
    grouped_df = filtered_df.groupby('子项名称')
    return grouped_df


# 使用Openpyxl读取指定的Excel模板，并填写内容
def fill_template(grouped_df, template_path, folder_name):
    filelist=[]
    for group_name, group_data in grouped_df:
        i = 1
        wb = load_workbook(template_path)
        ws = wb.active
        ws["O2"].value = group_name
        rows_count = len(group_data)
        print(group_name,rows_count)
        # 判断是否需要分组填写
        if rows_count <= 29:
            current_row = 4
            for index, row in group_data.iterrows():
                ws["A" + str(current_row)].value = row['设备名称']
                ws["C" + str(current_row)].value = row['电动机型号']
                ws["F" + str(current_row)].value = row['制造厂']
                ws["I" + str(current_row)].value = row['产品编号']
                ws["K" + str(current_row)].value = row['额定电压']
                ws["M" + str(current_row)].value = row['额定电流']
                ws["O" + str(current_row)].value = row['额定功率']
                ws["Q" + str(current_row)].value = ">10"
                current_row += 1
            output_path = folder_name + f"{group_name}.xlsx"
            wb.save(output_path)
            filelist.append(output_path)
        else:
            current_row = 4
            group_count = rows_count // 29 + (1 if rows_count % 29 != 0 else 0)
            for j in range(group_count):
                group_data_subset = group_data.iloc[j*29:min((j+1)*29, rows_count)]
                for index, row in group_data_subset.iterrows():
                    ws["A" + str(current_row)].value = row['设备名称']
                    ws["C" + str(current_row)].value = row['电动机型号']
                    ws["F" + str(current_row)].value = row['制造厂']
                    ws["I" + str(current_row)].value = row['产品编号']
                    ws["K" + str(current_row)].value = row['额定电压']
                    ws["M" + str(current_row)].value = row['额定电流']
                    ws["O" + str(current_row)].value = row['额定功率']
                    ws["Q" + str(current_row)].value = ">10"
                    current_row += 1
                output_path = folder_name + f"{group_name}_{i}.xlsx"
                wb.save(output_path)
                filelist.append(output_path)
                i += 1
                wb = load_workbook(template_path)
                ws = wb.active
                ws["O2"].value = group_name
                current_row = 4
            output_path = folder_name + f"{group_name}_{i}.xlsx"
            wb.save(output_path)
            filelist.append(output_path)
    return filelist





# 创建文件夹并保存文件
def create_folder_and_save_files(grouped_df,folder_name):
    # os.makedirs(folder_name, exist_ok=True)
    for group_name, group_data in grouped_df:
        group_folder = os.path.join(folder_name, group_name)
        # os.makedirs(group_folder, exist_ok=True)



# 使用LibreOffice的API将所有文件转换为PDF并合并
# 将文件转换为PDF
def convert_to_pdf(input_file, output_file):
    cmd = ["D:\Software\Libre Offices\program\soffice.com", "--headless", "--convert-to", "pdf", input_file, "--outdir", os.path.dirname(output_file)]
    subprocess.run(cmd)

# 合并PDF文件
def merge_pdfs(pdf_files, output_pdf):
    # 创建PdfFileMerger对象
    merger = PdfFileMerger()

    # 合并PDF文件
    for pdf_file in pdf_files:
        merger.append(pdf_file)

    # 保存合并后的PDF文件
    merger.write(output_pdf)
    merger.close()

    # 删除转换的PDF文件
    for pdf_file in pdf_files:
        os.remove(pdf_file)
    print("合并完成")


# 主函数
def main():
    # 读取Excel文件并转换为DataFrame
    file_path = "D:\Jobs\洛钼\调试记录\低压电机\电动机数据源.xlsx"
    sheet_name = "电机数据"
    df = read_excel_to_dataframe(file_path, sheet_name)

    # 筛选和分组
    grouped_df = filter_and_group(df)

    # 填写模板并保存文件
    folder_name = "D:\Jobs\洛钼\调试记录\低压电机\小电机\\"
    template_path = "D:\Jobs\洛钼\调试记录\低压电机\\3.2.3  100kW及以下低压电动机.xlsx"
    files = fill_template(grouped_df, template_path,folder_name)
    # 创建文件夹并保存文件
    # create_folder_and_save_files(grouped_df,folder_name)

    # 将文件夹内文件转换为PDF并合并
    for file in files:
        print(file)
        convert_to_pdf(pathlib.Path(file),folder_name)


    # convert_and_merge_to_pdf(folder_name)


if __name__ == "__main__":
    main()
