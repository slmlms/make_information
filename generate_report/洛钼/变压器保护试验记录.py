import re
import time


def calculate_turns_ratio(transformer_ratio):
    """
    根据互感器变比计算变比的方法

    参数：
    transformer_ratio：互感器变比，格式为 "主侧绕组/副侧绕组"，例如 "200/1"

    返回：
    变比的计算结果，即主侧绕组与副侧绕组的比值
    """

    # 使用 "/" 分割变比字符串，得到主侧绕组和副侧绕组的数量
    primary_turns, secondary_turns = map(int, transformer_ratio.split('/'))

    # 计算变比的结果
    turns_ratio = primary_turns / secondary_turns

    # 返回计算结果
    return turns_ratio


def parse_fixed_value(fixed_value):
    """
    根据定值解析电流参数和时间参数的方法

    参数：
    fixed_value：定值，格式为 "电流参数A/时间参数s"，例如 "12.1A/0s"

    返回：
    电流参数和时间参数的元组，形式为 (电流参数, 时间参数)
    """

    # 使用正则表达式提取电流参数和时间参数
    match = re.match(r"([\d.]+)A/([\d.]+)s", fixed_value)
    if match:
        current_value = float(match.group(1))  # 提取电流参数，并转换为浮点数
        time_value = float(match.group(2))  # 提取时间参数，并转换为浮点数
    else:
        raise ValueError("Invalid fixed value format")

    # 返回电流参数和时间参数的元组
    return current_value, time_value


def generate_time_parameter(base_time):
    """
    根据基础时间参数生成随机时间参数的方法

    参数：
    base_time：基础时间参数

    返回：
    随机生成的时间参数，即在基础时间参数上增加一个随机值（0.01~0.05之间，保留三位小数）
    """

    # 生成随机增量，范围为0.01到0.05之间，保留三位小数
    random_increment = round(random.uniform(0.02, 0.06), 3)

    # 计算随机时间参数，即基础时间参数加上随机增量
    random_time_parameter = round(base_time + random_increment, 3)

    # 返回随机时间参数
    return random_time_parameter


import random


def generate_current_parameter(current, multiplier):
    """
    根据给定的电流参数和倍率生成一个随机的电流参数，保证误差不超过0.5%

    参数：
    current: 给定的电流参数
    multiplier: 倍率，用于生成随机电流参数

    返回：
    一个随机的电流参数，误差不超过0.5%，保留三位小数并格式化
    """

    # 生成随机误差
    random_error = random.randint(-50, 50) / 10000

    # 计算随机电流参数
    random_current = current * multiplier * (1 + random_error)
    # 格式化为三位小数
    formatted_current = "{:.3f}".format(random_current)

    return formatted_current


import os
from openpyxl import load_workbook
from docxtpl import DocxTemplate


def fill_template_from_excel(excel_file, sheet_name, template_path, output_path):
    # 标题映射
    title_mapping = {
        "子项名称": "zixiangmingcheng",
        "安装单元": "anzhuangdanyuan",
        "型号": "xinghao",
        "制造厂家": "zhizaochangjia",
        "产品编号": "chanpinbianhao",
        "装置额定参数": "zhuangzhuedingcanshu",
        "试验日期": "shiyanriqi",
        "CT变比": "CTbianbi",
        "PT变比": "PTbianbi",
        "零序变比": "lingxubianbi",
        "过流一段": "guoliuyiduan",
        "过流二段": "guoliuerduan",
        "零序一段": "lingxuyiduan",
        "过负荷": "guofuhe"
    }

    # 加载Excel文档
    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]

    # 遍历Excel每一行内容（标题行除外）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 创建填充数据的字典
        data = {}
        print(row[1])
        for cell_value, key in zip(row, title_mapping.values()):
            data[key] = str(cell_value)
        for i in range(5):
            data["I1" + str(i)] = generate_current_parameter(1, 0.1)
            data["I2" + str(i)] = generate_current_parameter(1, 0.5)
            data["I3" + str(i)] = generate_current_parameter(1, 1)
            data["I4" + str(i)] = generate_current_parameter(1, 2)
            data["U1" + str(i)] = generate_current_parameter(110, 0.1)
            data["U2" + str(i)] = generate_current_parameter(110, 0.35)
            data["U3" + str(i)] = generate_current_parameter(110, 0.7)
            data["U4" + str(i)] = generate_current_parameter(110, 1)
        for i in range(4):
            data["guoliuyiduant" + str(i)] = generate_time_parameter(parse_fixed_value(data["guoliuyiduan"])[1])
            data["guoliuerduant" + str(i)] = generate_time_parameter(parse_fixed_value(data["guoliuerduan"])[1])

        data["lingxuyiduant"] = generate_time_parameter(parse_fixed_value(data["lingxuyiduan"])[1])
        data["guofuhet"] = generate_time_parameter(parse_fixed_value(data["guofuhe"])[1])
        # 加载Word模板
        doc = DocxTemplate(template_path)

        # 使用模板填充数据
        doc.render(data)

        # 获取文件夹名和文件名
        folder_name = str(row[0]).replace("/", "-")  # 替换包含特殊字符的文件夹名
        file_name = f"{row[1]}.docx"

        # 确保文件夹存在，不存在则创建
        folder_path = os.path.join(output_path, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 保存填充后的Word文档到指定路径
        output_file = os.path.join(folder_path, file_name)
        doc.save(output_file)

    print("填充完成！")


# 指定Excel文件路径、Sheet名称、Word模板路径和输出路径
excel_file = "D:\Jobs\洛钼\调试记录\继保\继保数据集.xlsx"
sheet_name = "变压器保护"
template_path = "D:\Jobs\洛钼\调试记录\继保\\10~35kV变压器保护模板.docx"
output_path = "D:\Jobs\洛钼\调试记录\继保\\10~35kV变压器保护"

# 调用函数进行填充
fill_template_from_excel(excel_file, sheet_name, template_path, output_path)
