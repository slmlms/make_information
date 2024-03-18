import os
import random

import openpyxl
from docxtpl import DocxTemplate
from openpyxl.reader.excel import load_workbook
from tqdm import tqdm

Jianyanpi_data_path = "D:\Jobs\洛钼\仪表和电信分部分项检验批\电信消防.xlsx"
Jianyanpi_save_path = "D:\Jobs\洛钼\仪表和电信分部分项检验批\电信消防\\"
Jianyanpi_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\智能建筑\检验批\\"
Fenxiang_tamepate_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\智能建筑\分项质量检查验收记录.docx"
Fenxiang_Baoyan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\智能建筑\分项报验申请表.docx"
Fenbu_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\智能建筑\分部工程质量检验评定记录.docx"
Fenbu_Baoyan_template_path = "D:\Documents\PycharmProjects\make_information\\resources\inspection_lot\智能建筑\分部报验申请表.docx"

# 给定一个Sheet名称，读取Excel表格，并返回一个Dataframe
import pandas as pd


def read_excel_to_dataframe(file_path, sheet_name):
    """
    从Excel文件中读取指定工作表到DataFrame中

    Args:
        file_path (str): Excel文件的路径
        sheet_name (str): 需要读取的工作表名称

    Returns:
        pandas.DataFrame: 读取的工作表数据作为DataFrame对象
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df


def get_template_file_path(sheet, row):
    template_file_path = Jianyanpi_template_path + sheet.cell(row=row, column=3).value + "检验批质量检验评定记录.docx"

    return template_file_path


def extract_filename_from_path(file_path):
    """
    从文件路径中提取文件名

    Args:
        file_path (str): 文件路径

    Returns:
        str: 文件名（不包含扩展名）
    """
    base_name = os.path.basename(file_path)  # 获取路径的基本名称（即文件名）
    filename_without_extension, extension = os.path.splitext(base_name)  # 分离文件名和扩展名
    return filename_without_extension


def jyp_dianqipeiguan(data):
    for i in range(1, 11):
        data["yigewan" + str(i)] = str(random.randint(4, 8)) + "D"  # 管子只有一个弯
        data["lianggewan" + str(i)] = str(random.randint(6, 9)) + "D"  # 管子有两个弯
    return data


# 生成检查单函数
def generate_inspection_batch(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        # 使用openpyxl库读取Excel表格
        wb = load_workbook(Jianyanpi_data_path)
        print("检验批")
        # 遍历表格中的每一个Sheet
        for sheet in wb:

            # 遍历每一行
            for row in tqdm(range(2, sheet.max_row + 1), desc=f'Processing Sheet: {sheet.title}'):
                if sheet.cell(row=row, column=1).value is None: continue
                try:
                    # 获取模板文件路径
                    template_file_path = get_template_file_path(sheet, row)
                    template_file = DocxTemplate(template_file_path)

                    # 获取数据
                    data = {
                        "Danweigongchengmingcheng": sheet.cell(row=row, column=1).value,
                        "Fenbugongchengmingcheng": sheet.cell(row=row, column=2).value,
                        "Fenxianggongchengmingcheng": sheet.cell(row=row, column=3).value,
                        "Jianyanchibuwei": sheet.cell(row=row, column=4).value,
                        # "Jianyanchirongliang": sheet.cell(row=row, column=5).value
                    }
                    if data["Fenxianggongchengmingcheng"] == "电气配管":
                        data = jyp_dianqipeiguan(data)
                    # 遍历字典并替换None为''
                    for key in data:
                        if data[key] is None:
                            data[key] = ''

                    # 构建保存路径
                    save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/" + data[
                        'Fenxianggongchengmingcheng'] + '/'
                    if not os.path.exists(os.path.dirname(save_path)):
                        os.makedirs(os.path.dirname(save_path))

                    # 渲染模板文件并保存
                    template_file.render(data)
                    template_file.save(
                        save_path + extract_filename_from_path(template_file_path) + data["Jianyanchibuwei"] + '.docx')

                except FileNotFoundError:
                    print(f"模板文件 {template_file_path} 不存在")
                except IndexError:
                    print(f"在行 {row} 的列中找不到值")
                except KeyError:
                    print(f"数据字典中缺少键（单元格值可能为空）")
                except Exception as e:
                    print(f"处理Sheet '{sheet.title}' 行 {row} 时发生错误: {str(e)}")

    except FileNotFoundError:
        print(f"找不到文件: {Jianyanpi_data_path}")
    except Exception as e:
        print(f"加载Excel文件时发生错误: {str(e)}")
    finally:
        # 如果需要的话，在这里关闭工作簿
        wb.close()


def generate_itemised_project(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        # 使用openpyxl库读取Excel表格
        wb = load_workbook(Jianyanpi_data_path)
        print("分项")
        for sheet in wb:
            # 打印Sheet名称
            print(sheet.title)

            try:
                df = read_excel_to_dataframe(Jianyanpi_data_path, sheet.title)

                unique_project_names = df['分项工程名称'].unique()

                for unique_project_name in tqdm(unique_project_names,
                                                desc=f'Processing Projects in Sheet: {sheet.title}'):
                    try:
                        tpl = DocxTemplate(Fenxiang_tamepate_path)  # 确保Fenxiang_tamepate_path变量已定义
                        tpl_baoyan = DocxTemplate(Fenxiang_Baoyan_template_path)
                        lists = []
                        data = {}

                        data["Danweigongchengmingcheng"] = sheet.title
                        project_name = df[df['分项工程名称'] == unique_project_name]['分部工程名称'].unique()
                        if len(project_name) > 0:
                            data["Fenbugongchengmingcheng"] = project_name[0]
                            data["Fenxianggongchengmingcheng"] = unique_project_name

                        part_names = df[df['分项工程名称'] == unique_project_name]['检验批部位'].unique()
                        data["size"] = len(part_names)

                        for i in range(len(part_names)):
                            list = {}
                            list["Xuhao"] = str(i + 1).zfill(2)
                            list["Jianyanchibuwei"] = part_names[i]
                            list["Shigongdanweipingding"] = "□优良    □合格"
                            list["Jianlidanweipingding"] = "□优良    □合格"
                            lists.append(list)
                        data["list"] = lists

                        save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/" + \
                                    data[
                                        'Fenxianggongchengmingcheng'] + '/'

                        if not os.path.exists(os.path.dirname(save_path)):
                            os.makedirs(os.path.dirname(save_path))
                        #
                        tpl.render(data)
                        tpl.save(save_path + "02" + data["Fenxianggongchengmingcheng"] + '分项质量检查验收记录.docx')
                        tpl_baoyan.render(data)
                        tpl_baoyan.save(
                            save_path + "01" + data["Fenxianggongchengmingcheng"] + '分项质量检查验收记录-报验表.docx')
                    except FileNotFoundError as fnfe:
                        print(f"模板文件 {Fenxiang_tamepate_path} 不存在. 错误信息: {fnfe}")
                    except KeyError as ke:
                        print(f"在DataFrame中找不到相应的列名. 错误信息: {ke}")
                    except IndexError as ie:
                        print(f"项目名称或部位名称的索引超出范围. 错误信息: {ie}")
                    except Exception as e:
                        print(f"处理Sheet '{sheet.title}' 时发生错误: {str(e)}")

            except ImportError as ime:
                print(f"未能正确导入或使用read_excel_to_dataframe函数. 错误信息: {ime}")
            except ValueError as ve:
                print(f"无法从Excel文件创建DataFrame. 错误信息: {ve}")
            except Exception as e:
                print(f"处理Sheet '{sheet.title}' 数据时发生错误: {str(e)}")

    except FileNotFoundError as fnfe:
        print(f"找不到Excel文件: {Jianyanpi_data_path}. 错误信息: {fnfe}")
    except openpyxl.exceptions.InvalidFileException as ife:
        print(f"Excel文件 {Jianyanpi_data_path} 格式无效. 错误信息: {ife}")
    except Exception as e:
        print(f"加载或处理Excel文件时发生错误: {str(e)}")


def generate_fenbu_project(Jianyanpi_data_path, Jianyanpi_save_path):
    try:
        # 使用openpyxl库读取Excel表格
        wb = load_workbook(Jianyanpi_data_path)
        print("分部")
        for sheet in wb:

            # 打印Sheet名称
            print(sheet.title)

            df = read_excel_to_dataframe(Jianyanpi_data_path, sheet.title)
            unique_project_names = df['分部工程名称'].unique()

            for unique_project_name in tqdm(unique_project_names, desc=f'Processing Projects in Sheet: {sheet.title}'):
                try:

                    data = {}

                    data['Danweigongchengmingcheng'] = sheet.title
                    data['Fenbugongchengmingcheng'] = unique_project_name
                    tpl = DocxTemplate(Fenbu_template_path)
                    tpl_Baoyan = DocxTemplate(Fenbu_Baoyan_template_path)

                    part_names = df[df['分部工程名称'] == unique_project_name]['分项工程名称'].unique()
                    data['Fenxiang_size'] = len(part_names)

                    lists = []
                    for i in range(len(part_names)):
                        list_dict = {}
                        list_dict['Xuhao'] = str(i + 1)
                        list_dict['Fenxianggongchengmingcheng'] = part_names[i]
                        list_dict['count_of_jianyanpi'] = str(
                            len(df[df['分项工程名称'] == part_names[i]]['检验批部位'].unique()))
                        list_dict["Shigongdanweipingding"] = "□优良    □合格"
                        list_dict["Jianlidanweipingding"] = "□优良    □合格"
                        lists.append(list_dict)

                    # 确保 lists 至少有 9 个元素
                    while len(lists) < 10:
                        lists.append({})

                    data['list'] = lists
                    save_path = Jianyanpi_save_path + sheet.title + "/" + data['Fenbugongchengmingcheng'] + "/"

                    if not os.path.exists(os.path.dirname(save_path)):
                        os.makedirs(os.path.dirname(save_path))

                    tpl.render(data)
                    tpl_Baoyan.render(data)

                    # 添加对保存路径和文件名的异常处理
                    try:
                        tpl.save(save_path + "02" + data['Fenbugongchengmingcheng'] + "分部工程质量检验评定记录.docx")
                        tpl_Baoyan.save(
                            save_path + "01" + data['Fenbugongchengmingcheng'] + "分部工程质量检验评定记录-报验表.docx")
                    except Exception as e:
                        print(f"保存Word文档时发生错误: {e}")

                except Exception as e:
                    print(f"处理{unique_project_name}时发生错误: {e}")

    except FileNotFoundError:
        print(f"无法找到Excel文件: {Jianyanpi_data_path}")
    except Exception as e:
        print(f"处理Excel或生成报告时发生未知错误: {e}")


# 调用方法生成检验批
generate_inspection_batch(Jianyanpi_data_path, Jianyanpi_save_path)
# 生成分项及报验表
generate_itemised_project(Jianyanpi_data_path, Jianyanpi_save_path)
# 生成分部及报验表
generate_fenbu_project(Jianyanpi_data_path, Jianyanpi_save_path)
